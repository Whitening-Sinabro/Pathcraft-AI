from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = (
    ROOT
    / "Docs"
    / "2026-08-26_SMOKIEZONE_HYDROSPHERE_BONESHATTER_3_29_SSF_TRACKER.xlsx"
)

NAVY = "1F4E78"
BLUE = "5B9BD5"
LIGHT_BLUE = "DDEBF7"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
YELLOW = "FFD966"
LIGHT_YELLOW = "FFF2CC"
RED = "C00000"
LIGHT_RED = "F4CCCC"
GRAY = "D9E1F2"
WHITE = "FFFFFF"
DARK = "1F1F1F"

THIN_GRAY = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def add_status_validation(ws, column_letter, start_row, end_row=500):
    validation = DataValidation(
        type="list",
        formula1='"미착수,진행,완료,보류,불필요"',
        allow_blank=True,
    )
    validation.error = "미착수/진행/완료/보류/불필요 중 하나를 선택하세요."
    validation.errorTitle = "상태 값 오류"
    validation.prompt = "진행 상태를 선택하세요."
    validation.promptTitle = "상태"
    ws.add_data_validation(validation)
    validation.add(f"{column_letter}{start_row}:{column_letter}{end_row}")

    green_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    yellow_fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    gray_fill = PatternFill("solid", fgColor="E7E6E6")
    ws.conditional_formatting.add(
        f"{column_letter}{start_row}:{column_letter}{end_row}",
        FormulaRule(
            formula=[f'{column_letter}{start_row}="완료"'],
            fill=green_fill,
        ),
    )
    ws.conditional_formatting.add(
        f"{column_letter}{start_row}:{column_letter}{end_row}",
        FormulaRule(
            formula=[f'{column_letter}{start_row}="진행"'],
            fill=blue_fill,
        ),
    )
    ws.conditional_formatting.add(
        f"{column_letter}{start_row}:{column_letter}{end_row}",
        FormulaRule(
            formula=[f'{column_letter}{start_row}="보류"'],
            fill=yellow_fill,
        ),
    )
    ws.conditional_formatting.add(
        f"{column_letter}{start_row}:{column_letter}{end_row}",
        FormulaRule(
            formula=[f'{column_letter}{start_row}="불필요"'],
            fill=gray_fill,
        ),
    )


def add_table_sheet(
    wb,
    name,
    title,
    headers,
    rows,
    widths,
    *,
    status_column=None,
    hyperlink_columns=(),
    tab_color=BLUE,
):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(1, 1, title)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.font = Font(color=WHITE, bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 27

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    note = ws.cell(
        2,
        1,
        "영문 원본명 기준 · 상태 셀은 드롭다운 · 필터/고정 행 적용 · 2026-08-26 검증",
    )
    note.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    note.font = Font(color=DARK, italic=True, size=10)
    note.alignment = Alignment(vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[3].height = 34

    data_rows = rows or [[""] * len(headers)]
    for row_index, row in enumerate(data_rows, 4):
        for col_index, value in enumerate(row, 1):
            cell = ws.cell(row_index, col_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFE")
            if col_index in hyperlink_columns and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.style = "Hyperlink"
        ws.row_dimensions[row_index].height = 44

    last_row = 3 + len(data_rows)
    last_col_letter = ws.cell(3, len(headers)).column_letter
    table = Table(
        displayName=f"Table_{name.replace('-', '_')}",
        ref=f"A3:{last_col_letter}{last_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col_letter}{last_row}"

    for index, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(3, index).column_letter].width = width

    if status_column:
        add_status_validation(ws, ws.cell(3, status_column).column_letter, 4)
        for row in range(4, last_row + 1):
            ws.cell(row, status_column).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    ws.auto_filter.ref = f"A3:{last_col_letter}{last_row}"
    return ws


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    start = wb.create_sheet("00_시작")
    start.sheet_view.showGridLines = False
    start.sheet_properties.tabColor = ORANGE
    start.merge_cells("A1:F1")
    start["A1"] = "PoE 3.29 SSF · Hydrosphere Boneshatter Berserker"
    start["A1"].fill = PatternFill("solid", fgColor=NAVY)
    start["A1"].font = Font(color=WHITE, bold=True, size=16)
    start["A1"].alignment = Alignment(vertical="center")
    start.row_dimensions[1].height = 32

    start.merge_cells("A2:F2")
    start["A2"] = (
        "smokie_777 메인 PoB + SSF 진행 영상 + Day 1/2/3 실제 PoB 교차검증"
    )
    start["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    start["A2"].font = Font(italic=True)

    summary_rows = [
        ("빌드", "Marauder → Berserker", "", "", "", ""),
        ("핵심 순서", "Sunder → 일반 Boneshatter → Warcry → Complex Trauma → 선택형 Crit", "", "", "", ""),
        ("초반 필수", "Blood Magic + Petrified Blood", "", "", "", ""),
        ("Complex Trauma", "레벨 90+, 방어도 15k+, 적합한 벨트, Sustainable Trauma 8", "", "", "", ""),
        ("Hydrosphere", "Battlemage's Cry 링크에는 실제로 젬 2개 사용", "", "", "", ""),
        ("폐기 세팅", "여러 레벨의 CWDT + Hydrosphere 실험은 불안정", "", "", "", ""),
        ("메인 영상", "PoE 3.29 League Starter", "https://www.youtube.com/watch?v=LwiJh2wc75w", "", "", ""),
        ("메인 PoB", "단계별 Skill Set과 Notes", "https://poe.ninja/poe1/pob/934b0", "", "", ""),
        ("SSF 진행", "아틀라스·타깃 파밍 순서", "https://www.youtube.com/watch?v=VpaneE2CVDs", "", "", ""),
    ]
    for row_index, row in enumerate(summary_rows, 4):
        for col_index, value in enumerate(row, 1):
            cell = start.cell(row_index, col_index, value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_index == 1:
                cell.fill = PatternFill("solid", fgColor=GRAY)
                cell.font = Font(bold=True)
            elif row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFE")
            if isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.style = "Hyperlink"
        start.row_dimensions[row_index].height = 34

    start["A15"] = "체크리스트 진행률"
    start["A15"].font = Font(bold=True, color=WHITE)
    start["A15"].fill = PatternFill("solid", fgColor=GREEN)
    start["B15"] = '=COUNTIF(\'06_체크리스트\'!A:A,"완료")'
    start["C15"] = "/"
    start["D15"] = "=COUNTA('06_체크리스트'!C4:C500)"
    for cell in start[15]:
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    start["B15"].number_format = "0"
    start["D15"].number_format = "0"

    start["A17"] = "버전 주의"
    start["A17"].font = Font(bold=True, color=WHITE)
    start["A17"].fill = PatternFill("solid", fgColor=RED)
    start.merge_cells("B17:F18")
    start["B17"] = (
        "영상은 3.29 자료지만 메인 PoB 934b0 내부 passive treeVersion은 3_28이다. "
        "3.29 PoB에서 변환 경고와 패시브 위치를 확인하고 Day 2/3 PoB를 후기 검증용으로 함께 사용한다."
    )
    start["B17"].alignment = Alignment(vertical="top", wrap_text=True)
    start["B17"].fill = PatternFill("solid", fgColor=LIGHT_RED)
    start["B17"].border = BORDER
    start["A18"].fill = PatternFill("solid", fgColor=RED)
    start["A18"].border = BORDER

    widths = [18, 65, 55, 8, 12, 12]
    for index, width in enumerate(widths, 1):
        start.column_dimensions[get_column_letter(index)].width = width
    start.freeze_panes = "A4"

    progression_rows = [
        ["미착수", "A. 캠페인", "레벨 1", "Ground Slam → Lv12 Sunder", "캠페인·도끼 갱신", "2차 전직과 흡수·추가 타격 패시브"],
        ["미착수", "B. 일반 Boneshatter", "2차 전직 완료", "4L Boneshatter, Blood Magic + Petrified Blood", "빠른 Despot/Ezomyte Axe, 생명력·저항", "T16 진입과 5링크"],
        ["미착수", "C. Warcry", "War Bringer와 필요한 젬·소켓", "Autoexert Intimidating + 수동 Battlemage/Hydrosphere×2", "Uber Lab, GCP", "함성 품질과 안정적 보스전"],
        ["미착수", "D. 벨트 확보", "Breach/Heist/Sirus 파밍 가능", "아직 일반 Boneshatter 유지 가능", "Soul Tether → Replica Soul Tether 또는 Burden", "마나 기반 전환 조건 충족"],
        ["미착수", "E. Complex Trauma", "Lv90+, Armour 15k+, 벨트, 마나 흡수", "Complex Trauma + Less Duration", "Vaal Axe, -mana cost 2곳", "Sustainable Trauma 8, 실전에서 10 미도달"],
        ["미착수", "F. 4 Stone", "일반 Soul Tether 이상 + 엔드게임 세팅", "메인 6L과 Warcry 6소켓", "보스·Voidstone", "4 Voidstone"],
        ["미착수", "G. Crit 선택", "Warcry crit of Infamy 투구", "Ruthless + Overexertion, 5 Warcries", "Mercenary 또는 Cyaxan's Ducat", "Crit cap 근접과 안정적인 5 Warcry"],
        ["미착수", "H. 최종", "Lv95+, 장비 최대화", "현재 완성 세팅 유지", "Simulacrum → Delirious Bloodline", "Delirious Bloodline"],
    ]
    add_table_sheet(
        wb,
        "01_단계별진행",
        "단계별 진행과 전환 조건",
        ["상태", "단계", "진입 조건", "젬·세팅 변화", "핵심 파밍", "단계 종료 조건"],
        progression_rows,
        [12, 22, 34, 54, 38, 38],
        status_column=1,
        tab_color=ORANGE,
    )

    gem_rows = [
        ["A", 1, "주력 3L", "초반 장비", "Ground Slam — Ruthless — Added Fire Damage", "사용 가능 레벨", "필수", "Momentum/Faster Attacks/Chance to Bleed는 비활성 대안"],
        ["A", 2, "주력 4L", "레벨 12+", "Sunder — Close Combat — Added Fire Damage — Fist of War", "사용 가능 레벨", "필수", "Rage/Melee Physical Damage는 비활성 대안"],
        ["A", 3, "이동", "초반", "Shield Charge — Faster Attacks — Momentum / Frostblink", "일반", "권장", "양손 도끼 뒤 Leap Slam"],
        ["A", 4, "예약·보조", "자유 소켓", "Blood and Sand / Intimidating Cry / Herald of Ash / Herald of Purity", "일반", "권장", "Clarity는 비활성"],
        ["B", 1, "본체 4L", "Body Armour", "Boneshatter — Close Combat — Fortify — Brutality", "일반", "필수", "5L에 Melee Physical Damage 추가"],
        ["B", 2, "예약", "자유 소켓", "Eternal Blessing — Pride / Vitality / Petrified Blood", "일반", "필수", "Blood Magic 패시브 필수"],
        ["B", 3, "Hydrosphere", "4L", "Hydrosphere — Hydrosphere — Urgent Orders — Battlemage's Cry", "Hydrosphere Lv1 가능", "필수", "두 복사본이 맞음; Battlemage's Cry 직접 사용"],
        ["B", 4, "CWDT", "4L", "CWDT Lv1 — Blood Rage Lv7 — Vulnerability Lv5 — Molten Shell Lv7", "표기 레벨 고정", "권장", "요구 레벨이 CWDT 한계를 넘지 않게 유지"],
        ["B", 5, "이동", "4소켓", "Leap Slam — Faster Attacks — Momentum / Frostblink", "일반", "필수", ""],
        ["C", 1, "본체 5L", "Body Armour", "Boneshatter — Close Combat — Brutality — Melee Physical Damage — Fortify", "20/20 우선 아님", "필수", ""],
        ["C", 2, "자동 함성", "2L", "Autoexertion — Intimidating Cry", "Intimidating Cry Lv1 가능", "필수", ""],
        ["C", 3, "수동 함성", "4L", "Hydrosphere ×2 — Urgent Orders — Battlemage's Cry", "함성 품질 우선", "필수", "우버랩 이후 보수적 전환"],
        ["C", 4, "Lab 임시", "예약", "Determination / Vitality", "일반", "상황", "Pride와 다른 생명력 예약을 빼고 Lab 안전 우선"],
        ["D", 1, "본체 6L", "Body Armour", "Complex Trauma — Close Combat — Brutality — Melee Physical — Less Duration — Fortify", "주력 20/30, Less Duration 20/20", "필수", "Fortify 패시브면 Fortify 대신 Ruthless 가능"],
        ["D", 2, "Warcry 6S", "Weapon", "Hydrosphere ×2 — Autoexertion — Battlemage's Cry — Intimidating Cry — Culling Strike", "Warcry 품질 우선", "필수", "전부 0 mana cost 목표"],
        ["D", 3, "오라", "Helmet 예시", "Pride / Flesh and Stone / Automation — Blood Rage", "Pride·F&S 21 가능", "필수", "Automation은 Blood Rage와 연결"],
        ["D", 4, "저주·방어", "Gloves 예시", "CWDT Lv20 — Vulnerability Lv20 — Lifetap / Vaal Molten Shell", "20/20 목표", "권장", "Vaal Molten Shell 직접 사용 가능"],
        ["D", 5, "이동", "Boots 예시", "Leap Slam — Faster Attacks / Frostblink / Enduring Cry", "일반", "필수", ""],
        ["E", 1, "Crit 본체 6L", "Body Armour", "Complex Trauma — Less Duration — Close Combat — Ruthless — Overexertion — Fortify", "20/20 목표", "조건부", "Warcry crit 투구 확보 후"],
        ["E", 2, "자동 함성 1", "Weapon", "Ancestral Cry — Power Charge On Critical — Intimidating Cry — Autoexertion — Infernal Cry", "Day 3 스냅샷", "조건부", "파워 차지 생성"],
        ["E", 3, "자동 함성 2", "Gloves 예시", "Battlemage's Cry — Hydrosphere ×2 — Autoexertion", "Day 3 스냅샷", "조건부", ""],
        ["E", 4, "직접 함성", "Boots 예시", "Rallying Cry — Urgent Orders", "20/20", "조건부", "다섯 번째 exerting Warcry"],
        ["폐기", 1, "실험", "여러 부위", "서로 다른 레벨 CWDT — Hydrosphere 여러 세트", "-", "사용 금지", "후속 검증에서 소켓 6개와 불규칙 발동 때문에 폐기"],
    ]
    add_table_sheet(
        wb,
        "02_젬세팅",
        "단계별 젬 세팅 — 메인 PoB와 Day 1·2·3 실제 PoB 대조",
        ["단계", "우선순위", "구분", "권장 슬롯", "링크·젬", "레벨·품질", "판정", "주의"],
        gem_rows,
        [9, 11, 18, 18, 70, 26, 14, 48],
        tab_color=GREEN,
    )

    atlas_rows = [
        [1, "Essence", "Contempt/Zeal 도끼, Insanity 갑옷", "좋은 빠른 도끼 + Onslaught 갑옷", "초기부터 유지"],
        [2, "Map sustain", "아틀라스 진행 안정화", "T16 유지", "Essence 사이에 확보"],
        [3, "Unwavering Vision", "초기 패시브 포인트 확보", "필요한 트리 완성", "오랫동안 유지 가능"],
        [4, "Packed with Energy", "일반 Boneshatter 초기 전투 강화", "기본 장비 완성", "Complex Trauma 뒤 공격 속도 재확인"],
        [5, "Huck / Heist", "Combat Ready와 Replica Soul Tether 청사진", "Blueprint 비축 또는 벨트 획득", "Complex Trauma 뒤 Huck 공격 속도 주의"],
        [6, "Searing Exarch", "GCP, Vaal Orb, +1 strike implicit", "Warcry 품질과 Atlas 진행", ""],
        [7, "Blight (2nd tree)", "Berserking anoint", "Clear + Sepia + Opalescent", "중간 목걸이에도 적용"],
        [8, "Harvest", "Amethyst Ring, Turquoise Amulet, Vaal Axe", "Chaos res·속성·생명력 개선", "Yellow Lifeforce 중심"],
        [9, "Breach", "Soul Tether + 방어 베이스", "Immortal Ambition Soul Tether", "Hive/Fortress 우선"],
        [10, "Heist/Sirus 분기", "Replica Soul Tether 또는 Burden of Truth", "최종 벨트", "Legion Corrupted Soul은 비추천"],
        [11, "Boss / 4 Stone", "Voidstone과 엔드게임", "4 Voidstone", "Soul Tether + Complex Trauma 뒤"],
        [12, "Simulacrum", "Delirious Bloodline", "최종 Bloodline", "Lv95+ 및 장비 최대화 뒤"],
    ]
    atlas = add_table_sheet(
        wb,
        "03_아틀라스",
        "SSF 아틀라스와 파밍 순서",
        ["순서", "콘텐츠", "목적", "중단·통과 조건", "주의"],
        atlas_rows,
        [10, 26, 46, 42, 46],
        tab_color=BLUE,
    )
    atlas["A17"] = "참고 트리"
    atlas["B17"] = "https://poeplanner.com/a/6ZSo"
    atlas["B17"].hyperlink = atlas["B17"].value
    atlas["B17"].style = "Hyperlink"

    crafting_rows = [
        ["Axe", "Despot > Ezomyte", "Vaal Axe", "Contempt/Zeal Essence → Allflame 선택 → 후기 recombination", "flat phys / %phys / attack speed", "Complex Trauma 전 Vaal Axe 강제 금지"],
        ["Helmet", "Life/Armour/Res, DEX/INT craft", "Warcry crit of Infamy", "Mercenary 확인 또는 Prefix lock + Cyaxan's Ducat", "+(2–3)% crit per exerting Warcry", "Crit 분기 전에는 일반 방어 투구가 낫다"],
        ["Body Armour", "1,000+ Armour, 5L", "4k+ Armour, +1 max res", "Essence of Insanity 반복", "Onslaught 6 sec when Hit", "자기 피해 생존 핵심"],
        ["Gloves", "Life/Res", "Attack speed, +1 strike", "일반 제작 + Eldritch implicit", "damage while leeching", "Breach 클리어 개선"],
        ["Boots", "20%+ movement speed", "Life/Res/DEX, cannot be chilled", "DEX Essence → suffix lock → Veiled Orb", "movement speed + cannot be chilled", ""],
        ["Amulet", "Turquoise, Life/Attributes", "Life + attributes, -mana cost", "Harvest/Recombine/Multimod", "Berserking anoint", "+1 phys gem은 자기 피해도 증가"],
        ["Ring 1", "Life/Res", "Amethyst, -mana cost", "Harvest Reforge Chaos", "Warcries cannot Exert Travel Skills", ""],
        ["Ring 2", "Life/Res", "Amethyst, -mana cost", "Harvest Reforge Chaos", "reduced Shock effect", "Mana Leech를 여기서 확보 가능"],
        ["Belt", "Rare life belt", "Replica Soul Tether/Burden", "Breach → Heist 또는 Sirus", "Corrupted Soul/혼합 방어", "일반 Soul Tether는 중간 단계"],
        ["Flasks", "Life + Quicksilver + 원소 플라스크", "피격 충전 + 자동 사용", "Alteration/Instilling", "bleed/CB, curse, armour, res", "Stormblood는 Crit 후기 선택"],
    ]
    add_table_sheet(
        wb,
        "04_장비제작",
        "장비 목표와 SSF 제작",
        ["부위", "초기 목표", "후기 목표", "제작 경로", "필요 옵션", "주의"],
        crafting_rows,
        [16, 30, 34, 50, 42, 46],
        tab_color=ORANGE,
    )

    farming_rows = [
        ["미착수", "5-link body", "일반 Boneshatter 링크", "초기 도박/Breach/일반 드롭", "5링크 확보", "4링크 유지"],
        ["미착수", "Essence of Insanity", "Onslaught 갑옷", "보라 Essence + Vaal Orb", "좋은 고방어 갑옷", "일반 고방어 갑옷"],
        ["미착수", "Berserking anoint", "공격 속도와 Rage", "Blight: Clear + Sepia + Opalescent", "성유 완료", "Deep Breaths는 공격 속도 과다일 때"],
        ["미착수", "Soul Tether", "마나 기반 전환", "Breach Hive/Fortress", "Immortal Ambition 확인", "계속 Petrified Blood"],
        ["미착수", "Replica Soul Tether", "선호 최종 벨트", "Grand Heist Blueprint", "벨트 드롭", "The Burden of Truth"],
        ["미착수", "The Burden of Truth", "Replica 대안", "Sirus", "벨트 드롭", "Replica Soul Tether"],
        ["미착수", "Mana Leech", "마나 기반 유지", "Axe/Gloves/Amulet/Ring", "한 줄 확보", "해당 옵션 장비 유지"],
        ["미착수", "-mana cost ×2", "Warcry/Hydrosphere 0 cost", "Ring/Amulet craft", "모든 관련 젬 0 cost", "Petrified Blood 단계 유지"],
        ["미착수", "Crit helmet", "Day 3 전환", "Strength Mercenary 또는 Cyaxan's Ducat", "of Infamy 옵션", "비치명타 메인 세팅 유지"],
        ["미착수", "Stormblood Sapphire", "Crit chill/freeze", "Velka", "선택 드롭", "일반 Sapphire Flask"],
        ["미착수", "Delirious Bloodline", "최종 방어", "Simulacrum", "Lv95+·장비 완성", "후기 목표로 연기"],
    ]
    add_table_sheet(
        wb,
        "05_타깃파밍",
        "SSF 타깃 파밍 — 시작·중단 조건과 대안",
        ["상태", "목표", "이유", "파밍처·방법", "중단 조건", "대안"],
        farming_rows,
        [12, 28, 36, 48, 34, 34],
        status_column=1,
        tab_color=GREEN,
    )

    checklist_tasks = [
        ("캠페인", "Lv12 Sunder 전환", "주력 젬 확인"),
        ("캠페인", "Crave the Slaughter", "1차 전직"),
        ("캠페인", "Aspect of Carnage", "2차 전직"),
        ("캠페인", "Brine King / Garukhan Pantheon", "동결·감전 대응"),
        ("일반 Boneshatter", "Blood Magic + Petrified Blood", "마나·overleech"),
        ("일반 Boneshatter", "Hydrosphere 젬 2개", "Battlemage's Cry용"),
        ("일반 Boneshatter", "빠른 Despot/Ezomyte Axe", "Trauma ramp"),
        ("초기 지도", "5-link + Life/Res/Armour", "T16 준비"),
        ("초기 지도", "Essence → Maps → Unwavering Vision", "Atlas 기반"),
        ("초기 지도", "Packed with Energy → Heist → Exarch", "진행 버프와 재료"),
        ("중기", "Berserking anoint", "Blight"),
        ("중기", "Amethyst Rings / Turquoise Amulet", "Harvest"),
        ("중기", "Immortal Ambition Soul Tether", "Breach"),
        ("Complex Trauma", "Lv90+", "전환 문턱"),
        ("Complex Trauma", "Armour 15,000+", "자기 피해"),
        ("Complex Trauma", "Mana Leech 1줄", "마나 유지"),
        ("Complex Trauma", "-mana cost 2곳", "Warcry/Hydrosphere 0 cost"),
        ("Complex Trauma", "Less Duration", "필수 링크"),
        ("Complex Trauma", "Sustainable Trauma 8", "PoB 확인"),
        ("Complex Trauma", "실전 overflow 검사", "10 도달 여부"),
        ("엔드게임", "Replica Soul Tether 또는 Burden", "최종 벨트"),
        ("엔드게임", "4 Voidstone", "진행"),
        ("Crit 선택", "Warcry crit of Infamy helmet", "젬 전환 전제"),
        ("최종", "Lv95+ / gear maxed", "Simulacrum 전제"),
        ("최종", "Delirious Bloodline", "마지막 목표"),
    ]
    checklist_rows = [["미착수", phase, task, evidence, "", ""] for phase, task, evidence in checklist_tasks]
    add_table_sheet(
        wb,
        "06_체크리스트",
        "실전 체크리스트",
        ["상태", "단계", "할 일", "확인 이유", "완료 날짜", "메모"],
        checklist_rows,
        [12, 24, 46, 38, 18, 44],
        status_column=1,
        tab_color=GREEN,
    )

    risk_rows = [
        ["Shrine", "Acceleration Shrine", "공격 속도 증가로 Trauma 10 overflow", "사용하지 않음", "일반 이동 속도로 진행"],
        ["Shrine", "Echoing Shrine", "Exert 작동 방해", "사용하지 않음", "해당 Shrine 회피"],
        ["Map", "Physical Reflect", "주력 물리 공격 반사", "회피", "리롤"],
        ["Map", "Cannot Leech", "생존·마나 유지 붕괴", "회피", "리롤"],
        ["Map", "Recovery 대폭 감소", "Trauma 자기 피해 유지 위험", "방어 충분할 때만", "리롤"],
        ["Map", "Armour/PDR 감소", "자기 피해와 물리 적중 위험", "여러 위험 옵션과 중첩 금지", "리롤"],
        ["Map", "-Maximum Resistance", "원소 최대 적중 감소", "고위험 콘텐츠에서 회피", "리롤"],
        ["Buff", "Huck / Haste", "Complex Trauma 공격 속도 변경", "PoB와 게임에서 스택 재확인", "Huck 제거"],
        ["Mechanic", "CWDT Hydrosphere 다중 세팅", "소켓 6개·불규칙 발동", "사용하지 않음", "Battlemage's Cry + Hydro×2"],
        ["Lab", "Pride + 생명력 예약 유지", "함정/Izaro 원샷 위험", "Pride → Determination, 예약 축소", "Lab 뒤 원복"],
    ]
    add_table_sheet(
        wb,
        "07_위험요소",
        "위험 요소와 대응",
        ["범주", "요소", "문제", "대응", "대안"],
        risk_rows,
        [16, 32, 48, 46, 40],
        tab_color=RED,
    )

    sources = [
        ["핵심", "Main guide", "PoE 3.29 SSF Boneshatter Hydrosphere Berserker League Starter", "https://www.youtube.com/watch?v=LwiJh2wc75w", "젬 약 12:52, 진행 전체"],
        ["핵심", "Main PoB", "934b0", "https://poe.ninja/poe1/pob/934b0", "4개 단계 Skill Set + Notes"],
        ["핵심", "SSF progression", "How to progress SMOKIEZONE optimally", "https://www.youtube.com/watch?v=VpaneE2CVDs", "Atlas/target farm"],
        ["참고", "Day 1", "SSF Bone Zerker Day 1", "https://www.youtube.com/watch?v=Hu59LoK94k0", "실제 초기 진행"],
        ["참고", "Day 1 PoB", "936d2", "https://poe.ninja/poe1/pob/936d2", "Day 1 젬"],
        ["핵심", "Day 2", "SSF Boneshatter Day 2", "https://www.youtube.com/watch?v=FTod4l3GlF8", "Replica/Complex Trauma"],
        ["핵심", "Day 2 PoB", "9386f", "https://poe.ninja/poe1/pob/9386f", "Day 2 젬"],
        ["후기", "Day 3", "4x Damage Crit Helmet", "https://www.youtube.com/watch?v=TYIsXTX7W7M", "Crit 분기"],
        ["후기", "Day 3 PoB", "93dc2", "https://poe.ninja/poe1/pob/93dc2", "Crit 젬·장비"],
        ["주의", "Mechanics test", "Is Hydrosphere + CWDT + Melee Splash Bait?", "https://www.youtube.com/watch?v=_e6srljMSpE", "CWDT 버전 폐기"],
        ["보조", "Campaign run", "3.28 5:53 Eater/Exarch overexplained", "https://www.youtube.com/watch?v=zrkmP90wSvY", "전체 주행"],
        ["후기", "Simulacrum", "Simulacrum Guide for Boneshatter", "https://www.youtube.com/watch?v=RNlLoWRDnis", "최종 준비"],
        ["성능", "Ubers", "SSF 10/10 Ubers", "https://www.youtube.com/watch?v=qxr0r0n68wg", "성능 증거"],
        ["검증", "Cyaxan's Ducat", "PoEDB", "https://poedb.tw/us/Cyaxans_Ducat", "영문 효과"],
        ["검증", "Warcry crit modifier", "PoE Wiki", "https://www.poewiki.net/wiki/Modifier:MercenaryModAddCritPerExert", "2–3% 범위"],
        ["검증", "Stormblood", "PoEDB", "https://poedb.tw/us/Stormblood", "옵션·Velka 드롭"],
    ]
    add_table_sheet(
        wb,
        "08_출처",
        "영문 원본 출처와 검증 범위",
        ["중요도", "종류", "제목·식별자", "링크", "확인 내용"],
        sources,
        [14, 22, 54, 70, 38],
        hyperlink_columns=(4,),
        tab_color=BLUE,
    )

    log_rows = [
        ["", "", "", "", "", "", "", "", ""],
    ]
    log = add_table_sheet(
        wb,
        "09_플레이로그",
        "플레이 로그 — 한 세션마다 한 줄",
        ["날짜", "레벨", "Atlas", "현재 단계", "주요 획득", "죽은 원인", "Trauma/APS", "다음 행동", "메모"],
        log_rows,
        [16, 10, 18, 22, 42, 40, 20, 44, 44],
        tab_color=YELLOW,
    )
    date_validation = DataValidation(type="date", allow_blank=True)
    log.add_data_validation(date_validation)
    date_validation.add("A4:A500")

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)


def validate_workbook():
    wb = load_workbook(OUTPUT, data_only=False)
    expected = [
        "00_시작",
        "01_단계별진행",
        "02_젬세팅",
        "03_아틀라스",
        "04_장비제작",
        "05_타깃파밍",
        "06_체크리스트",
        "07_위험요소",
        "08_출처",
        "09_플레이로그",
    ]
    assert wb.sheetnames == expected, wb.sheetnames
    assert wb["02_젬세팅"].max_row >= 25
    assert wb["06_체크리스트"].max_row >= 28
    assert wb["08_출처"]["D4"].hyperlink is not None
    assert OUTPUT.stat().st_size > 20_000
    print(f"Created and validated: {OUTPUT}")
    print(f"Size: {OUTPUT.stat().st_size:,} bytes")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    build_workbook()
    validate_workbook()
